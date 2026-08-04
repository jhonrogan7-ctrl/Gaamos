"""Read a menu spreadsheet into rows the build service can write.

Pure Python on purpose: no models, no Celery, no settings. The upload view
decides what to do with a rejection; this module only says what is wrong.

Two conventions carry real weight, and both come from the first real sheet:

* A placeholder dash is EMPTY. The sample sheet uses `—` in more than forty
  optional cells. Rejecting the file would mean hand-editing all of them;
  reading the dash as text would print `—` on a live menu.
* A note is the uncertainty signal. The assistant that fills the sheet is told
  to leave a note whenever a price is unreadable or a mapping is guessed, so
  `notes` non-empty is the only thing "needs checking" has to mean. Duplicate
  and absurd-price findings are written into the same field, which is why the
  red badge on a row can always say why it is red.
"""
from dataclasses import dataclass, field

import openpyxl

HEADERS = ('Category', 'Sub_Category', 'Item', 'Variant', 'Description',
           'Price', 'Image_Subject', 'Notes')

# Text a human or an assistant writes to mean "nothing here".
_PLACEHOLDERS = frozenset({'—', '–', '-', 'n/a', 'na', 'none', 'null', ''})

# A price above this is an order-of-magnitude typo, not a dish. Rs 100,000 is
# already far beyond anything a Pokhara card prints.
_ABSURD_PRICE = 100_000


@dataclass
class SheetRow:
    line: int
    category: str = ''
    sub_category: str = ''
    item: str = ''
    variant: str = ''
    description: str = ''
    price: int = None
    subject: str = ''
    notes: str = ''


@dataclass
class ParseResult:
    rows: list = field(default_factory=list)
    errors: list = field(default_factory=list)
    dashes: int = 0
    sheet_name: str = ''

    @property
    def ok(self):
        return not self.errors


def _sheet(wb):
    """The sheet named `Menu`, else the first one.

    An assistant asked for a workbook often names the sheet `Sheet1`. Rejecting
    the file for that is friction with no safety behind it, so the fallback is
    silent in behaviour and loud in the result: `ParseResult.sheet_name` says
    which sheet was read.
    """
    if 'Menu' in wb.sheetnames:
        return wb['Menu']
    return wb.worksheets[0]


def _clean(value):
    """-> (text, was_placeholder). A dash means empty, and we count it."""
    if value is None:
        return '', False
    text = str(value).strip()
    if text.lower() in _PLACEHOLDERS:
        return '', text != ''
    return text, False


def _price(value):
    """-> (int, error_message). Whole rupees only; no symbols, no decimals."""
    if value is None or str(value).strip() == '':
        return None, 'Price is required'
    if isinstance(value, bool):
        return None, 'Price is not a whole number'
    if isinstance(value, int):
        return value, ''
    if isinstance(value, float):
        # 250.0 written by a spreadsheet is a whole number; 12.5 is not.
        return (int(value), '') if value.is_integer() else \
            (None, 'Price is not a whole number')
    text = str(value).strip()
    if text.isdigit():
        return int(text), ''
    return None, 'Price is not a whole number'


def parse(fileobj):
    """Read a workbook. -> ParseResult.

    Structural faults reject the file; content faults annotate the row. The
    split matters because the sheet is machine-produced: a structural fault is
    the same fault on every row, and a content fault is one row a human should
    look at.
    """
    result = ParseResult()
    try:
        wb = openpyxl.load_workbook(fileobj, data_only=True, read_only=True)
    except Exception as exc:                       # noqa: BLE001 — any openpyxl failure
        result.errors.append((f'The file could not be read as .xlsx ({exc})', []))
        return result

    ws = _sheet(wb)
    result.sheet_name = ws.title
    grid = list(ws.iter_rows(values_only=True))
    if not grid:
        result.errors.append(('The sheet is empty', []))
        return result

    header = [(_clean(c)[0]) for c in grid[0][:len(HEADERS)]]
    if tuple(header) != HEADERS:
        result.errors.append((
            'The header row must be exactly: ' + ' | '.join(HEADERS), []))
        return result

    body = [r for r in grid[1:] if any(_clean(c)[0] for c in r[:len(HEADERS)])]
    if not body:
        result.errors.append(('The sheet has no data rows', []))
        return result

    problems = {}          # message -> [line numbers], so errors arrive grouped
    rows = []
    for offset, raw in enumerate(body):
        line = offset + 2                                   # 1-based, header is 1
        cells = list(raw[:len(HEADERS)]) + [None] * (len(HEADERS) - len(raw))
        values = []
        for cell in cells:
            text, was_dash = _clean(cell)
            result.dashes += 1 if was_dash else 0
            values.append(text)

        category, sub_category, item, variant, description, _, subject, notes = values
        price, price_error = _price(cells[5])

        row_problems = []
        if not category:
            row_problems.append('Category is required')
        if not item:
            row_problems.append('Item is required')
        if not subject:
            row_problems.append('Image_Subject is required')
        if price_error:
            row_problems.append(price_error)
        if row_problems:
            for message in row_problems:
                problems.setdefault(message, []).append(line)
            continue

        rows.append(SheetRow(
            line=line, category=category, sub_category=sub_category, item=item,
            variant=variant, description=description, price=price,
            subject=subject, notes=notes))

    if problems:
        result.errors = sorted(problems.items())
        return result

    _annotate(rows)
    result.rows = rows
    return result


def _note(row, text):
    row.notes = f'{row.notes} · {text}' if row.notes else text


def _annotate(rows):
    """Write content findings into `notes` — the one field review reads."""
    seen = {}
    for row in rows:
        key = (row.category.lower(), row.sub_category.lower(),
               row.item.lower(), row.variant.lower())
        first = seen.get(key)
        if first is None:
            seen[key] = row.line
        else:
            _note(row, f'Duplicate of row {first}')
        if row.price == 0:
            _note(row, 'Price is 0 — confirm this dish is free')
        elif row.price > _ABSURD_PRICE:
            _note(row, f'Price {row.price} is unusually high — check for a typo')
